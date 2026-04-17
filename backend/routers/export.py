"""
Export router — generate downloadable Excel reports from ingested filings.

  GET  /export/{doc_id}/excel  →  download an Excel workbook with metrics + key passages
"""

from __future__ import annotations

import io
import json
import os
from datetime import datetime

from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from backend.services import vector_store

router = APIRouter(prefix="/export", tags=["export"])


@router.get("/{doc_id}/excel")
def export_excel(doc_id: str):
    """
    Generate and stream an Excel workbook for a single ingested filing.
    Contains sheets for: Overview, Risk Factors, MD&A, Financial Statements.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is required: pip install openpyxl"
        )

    doc_meta = vector_store.get_document_metadata(doc_id)
    if not doc_meta:
        raise HTTPException(status_code=404, detail=f"Document '{doc_id}' not found.")

    ticker = doc_meta.get("ticker", "UNKNOWN")
    fiscal_year = doc_meta.get("fiscal_year", "")
    filing_type = doc_meta.get("filing_type", "10-K")
    company = doc_meta.get("company_name", ticker)

    # Try to load cached metrics
    cache_path = os.path.join("./chroma_db/_metrics_cache", f"{doc_id}.json")
    metrics: dict = {}
    if os.path.exists(cache_path):
        try:
            with open(cache_path) as f:
                metrics = json.load(f)
        except Exception:
            pass

    # Pull section-specific chunks
    from backend.services import embedder
    sections_data: dict[str, list[str]] = {}
    for section, query in [
        ("Risk Factors", "key risks material uncertainties regulatory compliance"),
        ("MD&A", "management discussion analysis operating results revenue growth"),
        ("Financial Statements", "revenue net income earnings per share balance sheet cash flow"),
    ]:
        q_emb = embedder.embed_query(query)
        children = vector_store.query_children(doc_id, q_emb, top_k=8)
        sections_data[section] = [c["text"] for c in children]

    # ── Build workbook ─────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ── Styles ─────────────────────────────────────────────────────────────────
    HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
    HEADER_FILL  = PatternFill(fill_type="solid", fgColor="1B4F72")
    TITLE_FONT   = Font(name="Calibri", bold=True, size=14, color="1B4F72")
    LABEL_FONT   = Font(name="Calibri", bold=True, size=11)
    WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
    CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

    def style_header_row(ws, row_num: int, cols: int) -> None:
        for col in range(1, cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER_ALIGN

    # ── Sheet 1: Overview ─────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Overview"
    ws1.column_dimensions["A"].width = 28
    ws1.column_dimensions["B"].width = 45

    ws1["A1"] = f"FinLens Report: {ticker} {filing_type} {fiscal_year}"
    ws1["A1"].font = TITLE_FONT
    ws1.merge_cells("A1:B1")
    ws1["A1"].alignment = CENTER_ALIGN

    ws1["A2"] = f"Company: {company}"
    ws1["A3"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws1["A4"] = f"Doc ID: {doc_id}"
    ws1["A5"] = f"Child Chunks: {doc_meta.get('child_count', 0)}"
    ws1["A6"] = f"Table Chunks: {doc_meta.get('table_count', 0)}"

    row = 8
    ws1.cell(row=row, column=1, value="Metric")
    ws1.cell(row=row, column=2, value="Value")
    style_header_row(ws1, row, 2)
    row += 1

    metric_display = [
        ("Revenue", _fmt_metric(metrics.get("revenue"), metrics.get("revenue_unit"))),
        ("Net Income", _fmt_metric(metrics.get("net_income"), metrics.get("net_income_unit"))),
        ("EPS (Diluted)", f"${metrics['eps_diluted']:.2f}" if metrics.get("eps_diluted") else "N/A"),
        ("Gross Margin %", f"{metrics['gross_margin_pct']:.1f}%" if metrics.get("gross_margin_pct") else "N/A"),
        ("Operating Margin %", f"{metrics['operating_margin_pct']:.1f}%" if metrics.get("operating_margin_pct") else "N/A"),
        ("Total Assets", _fmt_metric(metrics.get("total_assets"), metrics.get("total_assets_unit"))),
        ("Cash & Equivalents", _fmt_metric(metrics.get("cash_and_equivalents"), metrics.get("cash_unit"))),
        ("Long-Term Debt", _fmt_metric(metrics.get("long_term_debt"), metrics.get("debt_unit"))),
        ("R&D Expense", _fmt_metric(metrics.get("r_and_d_expense"), metrics.get("r_and_d_unit"))),
        ("Employees", f"{int(metrics['employees']):,}" if metrics.get("employees") else "N/A"),
    ]

    for label, value in metric_display:
        ws1.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws1.cell(row=row, column=2, value=value)
        row += 1

    # Key risks
    if metrics.get("key_risks"):
        row += 1
        ws1.cell(row=row, column=1, value="Key Risks Identified").font = LABEL_FONT
        row += 1
        for risk in metrics["key_risks"][:5]:
            ws1.cell(row=row, column=1, value="•")
            ws1.cell(row=row, column=2, value=risk)
            ws1.cell(row=row, column=2).alignment = WRAP_ALIGN
            row += 1

    # ── Section sheets ────────────────────────────────────────────────────────
    for sheet_name, chunks in sections_data.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 90

        ws.cell(row=1, column=1, value="#").font = HEADER_FONT
        ws.cell(row=1, column=1).fill = HEADER_FILL
        ws.cell(row=1, column=2, value=f"Excerpt — {sheet_name}").font = HEADER_FONT
        ws.cell(row=1, column=2).fill = HEADER_FILL
        ws.cell(row=1, column=2).alignment = CENTER_ALIGN

        for i, chunk in enumerate(chunks, start=2):
            ws.cell(row=i, column=1, value=i - 1)
            cell = ws.cell(row=i, column=2, value=chunk[:1000])
            cell.alignment = WRAP_ALIGN
            ws.row_dimensions[i].height = 60

    # ── Serialize and stream ──────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"FinLens_{ticker}_{filing_type}_{fiscal_year}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _fmt_metric(value: float | None, unit: str | None) -> str:
    if value is None:
        return "N/A"
    unit_str = f" {unit}" if unit else ""
    return f"{value:,.1f}{unit_str}"
